from django.forms import formset_factory
from django.shortcuts import render, redirect, get_object_or_404
from django.http import JsonResponse, HttpResponseRedirect, HttpResponse
from django.urls import reverse
from django.views.generic import ListView
from django.core.exceptions import ValidationError
from django.db.models import Count, Q, Sum
from django.contrib.auth.decorators import login_required, user_passes_test
from django.contrib.auth import login, logout, authenticate
from django.contrib import messages
from django.forms import modelformset_factory, formset_factory, inlineformset_factory
from django.db import transaction
from django.utils import timezone
from datetime import datetime, time as dt_time, timedelta
from django.db.models.functions import TruncMonth
from io import BytesIO
import csv

from .models import Cliente, Empleado, Servicio, Vehiculo, Reparacion, Agenda, Registro
from .forms import ClienteForm, VehiculoForm, EmpleadoForm, ServicioForm, ReparacionForm, CitaForm
from .decorators import es_jefe, es_encargado, puede_gestionar_empleados, puede_gestionar_servicios

# Importaciones para la API REST
from rest_framework import generics, viewsets, status
from rest_framework.response import Response
from rest_framework.decorators import action, api_view, permission_classes
from rest_framework.permissions import IsAuthenticated, IsAdminUser, AllowAny
from rest_framework.authentication import SessionAuthentication
from .serializers import *

# Funciones de ayuda para permisos
def es_jefe_o_encargado(user):
    """Verifica si el usuario es jefe o encargado"""
    return user.is_authenticated and (user.is_staff or hasattr(user, 'profile') and user.profile.es_empleado)

# ========== AUTENTICACIÓN Y DASHBOARD BÁSICOS ==========

def login_view(request):
    if request.method == 'POST':
        username = request.POST.get('username')
        password = request.POST.get('password')
        user = authenticate(request, username=username, password=password)
        if user is not None:
            login(request, user)
            messages.success(request, 'Inicio de sesión exitoso.')
            return redirect('inicio')
        messages.error(request, 'Usuario o contraseña incorrectos.')
    return render(request, 'auth/login.html')


@login_required
def logout_view(request):
    logout(request)
    messages.info(request, 'Has cerrado sesión.')
    return redirect('login')


@login_required
def perfil_view(request):
    return render(request, 'auth/perfil.html', {'user': request.user})


@login_required
def inicio(request):
    # Redirigir a panel del jefe si corresponde
    if es_jefe(request.user):
        return redirect('dashboard_jefe')
    total_clientes = Cliente.objects.count()
    total_empleados = Empleado.objects.count()
    total_servicios = Servicio.objects.count()
    total_vehiculos = Vehiculo.objects.count()
    reparaciones_pendientes = Reparacion.objects.filter(estado_reparacion='en_progreso').count()
    citas_hoy = Agenda.objects.filter(fecha=timezone.now().date()).count()
    context = {
        'total_clientes': total_clientes,
        'total_empleados': total_empleados,
        'total_servicios': total_servicios,
        'total_vehiculos': total_vehiculos,
        'reparaciones_pendientes': reparaciones_pendientes,
        'citas_hoy': citas_hoy,
    }
    return render(request, 'inicio.html', context)


# Vista temporal para otras rutas aún no implementadas
@login_required
def not_implemented_view(request, *args, **kwargs):
    messages.info(request, 'Funcionalidad en desarrollo.')
    return redirect('inicio')

# Alias temporales para dashboards aún no implementados
dashboard_encargado = not_implemented_view


@login_required
def buscar_clientes(request):
    q = request.GET.get('q', '').strip()
    clientes = Cliente.objects.all()
    if q:
        clientes = clientes.filter(
            Q(nombre__icontains=q) |
            Q(apellido__icontains=q) |
            Q(correo_electronico__icontains=q) |
            Q(telefono__icontains=q)
        )
    results = []
    for c in clientes[:10]:
        results.append({
            'id': c.id,
            'nombre': c.nombre,
            'apellido': c.apellido,
            'telefono': c.telefono,
            'correo_electronico': c.correo_electronico,
            # Campo alias para compatibilidad con template de vehículos
            'cedula': c.telefono or ''
        })
    # Devolver en dos formatos por compatibilidad con distintos JS en templates
    return JsonResponse({'results': results, 'clientes': results})

# ========== DASHBOARD JEFE ==========

@login_required
def dashboard_jefe(request):
    hoy = timezone.now().date()

    total_clientes = Cliente.objects.count()
    total_vehiculos = Vehiculo.objects.count()
    total_reparaciones = Reparacion.objects.count()
    total_servicios = Servicio.objects.count()
    reparaciones_pendientes = Reparacion.objects.filter(
        estado_reparacion__in=['pendiente', 'en_progreso', 'en_espera', 'revision']
    ).count()
    reparaciones_completadas = Reparacion.objects.filter(estado_reparacion='completada').count()
    citas_hoy_count = Agenda.objects.filter(fecha=hoy).count()
    clientes_nuevos_este_mes = Cliente.objects.filter(
        fecha_registro__year=hoy.year,
        fecha_registro__month=hoy.month
    ).count()

    # Vehículos con más reparaciones en últimos 30 días
    desde = hoy - timedelta(days=30)
    vehiculos_frecuentes = Vehiculo.objects.annotate(
        num_reparaciones=Count('reparaciones', filter=Q(reparaciones__fecha_ingreso__gte=desde))
    ).order_by('-num_reparaciones')[:5]

    # Reparaciones por estado (para gráfico)
    estado_map = {
        'pendiente': 'Pendiente',
        'en_progreso': 'En Proceso',
        'en_espera': 'En Espera',
        'revision': 'Para Revisión',
        'completada': 'Completada',
        'cancelada': 'Cancelada',
    }
    rep_estados_qs = Reparacion.objects.values('estado_reparacion').annotate(total=Count('id'))
    reparaciones_por_estado = [
        {
            'estado': estado_map.get(item['estado_reparacion'], item['estado_reparacion']),
            'total': item['total']
        }
        for item in rep_estados_qs
    ]

    # Ingresos mensuales (suma de costo del servicio por mes) - últimos 6 meses
    ingresos_qs = (Reparacion.objects
                   .annotate(m=TruncMonth('fecha_ingreso'))
                   .values('m')
                   .annotate(total=Sum('servicio__costo'))
                   .order_by('m'))
    ingresos_totales_agg = Reparacion.objects.aggregate(total=Sum('servicio__costo'))
    ingresos_totales = float(ingresos_totales_agg['total'] or 0)
    meses_all = [item['m'].strftime('%b %Y') if item['m'] else '' for item in ingresos_qs]
    ingresos_all = [float(item['total']) if item['total'] is not None else 0.0 for item in ingresos_qs]
    meses = meses_all[-6:]
    ingresos = ingresos_all[-6:]
    total_ingresos_mensuales = sum(ingresos) if ingresos else 0.0
    promedio_mensual = (total_ingresos_mensuales / len(ingresos)) if ingresos else None

    # Tiempo promedio de reparación (en días) para completadas
    completadas = Reparacion.objects.filter(fecha_salida__isnull=False)
    if completadas.exists():
        duraciones = [(r.fecha_salida - r.fecha_ingreso).days for r in completadas]
        avg_days = sum(duraciones) / len(duraciones) if duraciones else 0
        # Representar como timedelta
        from datetime import timedelta as _td
        tiempo_promedio = _td(days=int(round(avg_days)))
    else:
        tiempo_promedio = None

    # Listas para secciones
    reparaciones_recientes = Reparacion.objects.select_related('vehiculo', 'servicio').order_by('-fecha_ingreso')[:10]
    citas_proximas = Agenda.objects.select_related('cliente', 'servicio').filter(fecha__gte=hoy).order_by('fecha', 'hora')[:10]

    # Empleados destacados por registros (últimos 30 días)
    top_registros = (Registro.objects
                     .filter(fecha__gte=desde)
                     .values('empleado__id', 'empleado__nombre', 'empleado__puesto')
                     .annotate(num_reparaciones=Count('id'))
                     .order_by('-num_reparaciones')[:4])
    empleados_destacados = [
        {
            'nombre': r['empleado__nombre'],
            'puesto': r['empleado__puesto'],
            'num_reparaciones': r['num_reparaciones'],
        }
        for r in top_registros
    ]

    context = {
        'titulo': 'Panel del Jefe',
        'hoy': timezone.now(),
        'total_clientes': total_clientes,
        'total_vehiculos': total_vehiculos,
        'total_servicios': total_servicios,
        'total_reparaciones': total_reparaciones,
        'reparaciones_pendientes': reparaciones_pendientes,
        'reparaciones_completadas': reparaciones_completadas,
        'citas_hoy_count': citas_hoy_count,
        'clientes_nuevos_este_mes': clientes_nuevos_este_mes,
        'vehiculos_frecuentes': vehiculos_frecuentes,
        'reparaciones_por_estado': reparaciones_por_estado,
        'ingresos_mensuales': bool(ingresos),
        'meses': meses,
        'ingresos': ingresos,
        'ingresos_totales': ingresos_totales,
        'promedio_mensual': promedio_mensual,
        'total_ingresos_mensuales': total_ingresos_mensuales,
        'reparaciones_recientes': reparaciones_recientes,
        'citas_proximas': citas_proximas,
        'empleados_destacados': empleados_destacados,
        'tiempo_promedio': tiempo_promedio,
    }
    return render(request, 'gestion/dashboard_jefe.html', context)

# ========== DASHBOARD REPARACIONES Y CRUD ==========

@login_required
def dashboard_reparaciones(request):
    hoy = timezone.now()

    total_reparaciones = Reparacion.objects.count()
    reparaciones_completadas = Reparacion.objects.filter(estado_reparacion='completada').count()
    reparaciones_en_progreso = Reparacion.objects.filter(estado_reparacion='en_progreso').count()
    reparaciones_pendientes = Reparacion.objects.filter(estado_reparacion='pendiente').count()
    reparaciones_en_espera = Reparacion.objects.filter(estado_reparacion='en_espera').count()
    reparaciones_revision = Reparacion.objects.filter(estado_reparacion='revision').count()
    reparaciones_canceladas = Reparacion.objects.filter(estado_reparacion='cancelada').count()

    # Dict para el gráfico del template
    reparaciones_por_estado = {
        'Completada': reparaciones_completadas,
        'En_Progreso': reparaciones_en_progreso,
        'Pendiente': reparaciones_pendientes,
        'En_Espera': reparaciones_en_espera,
        'Lista_para_Revisión': reparaciones_revision,
        'Cancelada': reparaciones_canceladas,
    }

    # Servicios más solicitados
    servicios_qs = (Reparacion.objects
                    .values('servicio__nombre_servicio')
                    .annotate(total=Count('id'))
                    .order_by('-total')[:5])
    servicios_mas_solicitados = [
        {'nombre': s['servicio__nombre_servicio'], 'total': s['total']}
        for s in servicios_qs
    ]

    ultimas_reparaciones = (Reparacion.objects
                            .select_related('vehiculo', 'servicio', 'vehiculo__cliente')
                            .order_by('-fecha_ingreso')[:10])

    context = {
        'hoy': hoy,
        'total_reparaciones': total_reparaciones,
        'reparaciones_completadas': reparaciones_completadas,
        'reparaciones_en_progreso': reparaciones_en_progreso,
        'reparaciones_pendientes': reparaciones_pendientes,
        'reparaciones_por_estado': reparaciones_por_estado,
        'servicios_mas_solicitados': servicios_mas_solicitados,
        'ultimas_reparaciones': ultimas_reparaciones,
    }
    return render(request, 'gestion/dashboard_reparaciones.html', context)


@login_required
def crear_reparacion(request):
    titulo = 'Nueva Reparación'
    if request.method == 'POST':
        form = ReparacionForm(request.POST)
        if form.is_valid():
            # Crear la reparación pero no guardar aún
            reparacion = form.save(commit=False)
            # Asegurarse de que los campos requeridos tengan valores por defecto si no se proporcionan
            if not reparacion.condicion_vehiculo:
                reparacion.condicion_vehiculo = 'regular'
            if not reparacion.estado_reparacion:
                reparacion.estado_reparacion = 'pendiente'
            # Guardar la reparación con los valores por defecto si es necesario
            reparacion.save()
            messages.success(request, 'Reparación creada correctamente.')
            return redirect('dashboard_reparaciones')
        else:
            # Si el formulario no es válido, mostrar errores en la consola para depuración
            print("Errores del formulario:", form.errors)
    else:
        # Inicializar el formulario con valores por defecto
        form = ReparacionForm(initial={
            'condicion_vehiculo': 'regular',
            'estado_reparacion': 'pendiente'
        })
    return render(request, 'gestion/reparacion_form.html', {'form': form, 'titulo': titulo})


@login_required
def editar_reparacion(request, pk):
    reparacion = get_object_or_404(Reparacion, pk=pk)
    titulo = 'Editar Reparación'
    if request.method == 'POST':
        form = ReparacionForm(request.POST, instance=reparacion)
        if form.is_valid():
            form.save()
            messages.success(request, 'Reparación actualizada correctamente.')
            return redirect('dashboard_reparaciones')
    else:
        form = ReparacionForm(instance=reparacion)
    return render(request, 'gestion/reparacion_form.html', {'form': form, 'titulo': titulo})


@login_required
def eliminar_reparacion(request, pk):
    reparacion = get_object_or_404(Reparacion, pk=pk)
    if request.method == 'POST':
        reparacion.delete()
        messages.success(request, 'Reparación eliminada correctamente.')
        return redirect('dashboard_reparaciones')
    return render(request, 'gestion/reparacion_confirm_delete.html', {'reparacion': reparacion})

# ========== REPORTES ==========

@login_required
def reportes_ingresos(request):
    hoy = timezone.now()
    fecha_desde_str = request.GET.get('fecha_desde')
    fecha_hasta_str = request.GET.get('fecha_hasta')

    reparaciones = Reparacion.objects.all()
    fecha_desde = None
    fecha_hasta = None
    if fecha_desde_str:
        try:
            from datetime import datetime as _dt
            fecha_desde = _dt.strptime(fecha_desde_str, '%Y-%m-%d').date()
            reparaciones = reparaciones.filter(fecha_ingreso__date__gte=fecha_desde)
        except Exception:
            fecha_desde = None
    if fecha_hasta_str:
        try:
            from datetime import datetime as _dt
            fecha_hasta = _dt.strptime(fecha_hasta_str, '%Y-%m-%d').date()
            reparaciones = reparaciones.filter(fecha_ingreso__date__lte=fecha_hasta)
        except Exception:
            fecha_hasta = None

    ingresos_qs = (reparaciones
                   .annotate(m=TruncMonth('fecha_ingreso'))
                   .values('m')
                   .annotate(total=Sum('servicio__costo'), cantidad=Count('id'))
                   .order_by('m'))

    meses_all = [item['m'].strftime('%b %Y') if item['m'] else '' for item in ingresos_qs]
    ingresos_all = [float(item['total']) if item['total'] is not None else 0.0 for item in ingresos_qs]
    cantidades_all = [item['cantidad'] for item in ingresos_qs]

    # Si no hay filtros, mostrar últimos 12 meses; si hay filtros, mostrar todo el rango
    if not fecha_desde_str and not fecha_hasta_str:
        meses = meses_all[-12:]
        ingresos = ingresos_all[-12:]
        cantidades = cantidades_all[-12:]
    else:
        meses = meses_all
        ingresos = ingresos_all
        cantidades = cantidades_all

    ingresos_totales = sum(ingresos) if ingresos else 0.0
    promedio_mensual = (ingresos_totales / len(ingresos)) if ingresos else 0.0

    detalles = [
        {'mes': meses[i], 'total': ingresos[i], 'cantidad': cantidades[i]}
        for i in range(len(meses))
    ]

    context = {
        'titulo': 'Reporte de Ingresos',
        'hoy': hoy,
        'meses': meses,
        'ingresos': ingresos,
        'ingresos_totales': ingresos_totales,
        'promedio_mensual': promedio_mensual,
        'detalles': detalles,
        'fecha_desde': fecha_desde,
        'fecha_hasta': fecha_hasta,
    }
    return render(request, 'gestion/reportes_ingresos.html', context)

# Exportar ingresos a Excel con gráfico (xlsxwriter u openpyxl). Fallback a CSV si ninguna está instalada.
def exportar_ingresos_excel(request):
    fecha_desde_str = request.GET.get('fecha_desde')
    fecha_hasta_str = request.GET.get('fecha_hasta')
    reparaciones = Reparacion.objects.all()
    
    # Aplicar filtros de fecha si existen
    if fecha_desde_str:
        try:
            fecha_desde = datetime.strptime(fecha_desde_str, '%Y-%m-%d').date()
            reparaciones = reparaciones.filter(fecha_ingreso__date__gte=fecha_desde)
        except ValueError:
            pass
    
    if fecha_hasta_str:
        try:
            fecha_hasta = datetime.strptime(fecha_hasta_str, '%Y-%m-%d').date()
            reparaciones = reparaciones.filter(fecha_ingreso__date__lte=fecha_hasta)
        except ValueError:
            pass

    # Obtener datos para el reporte
    ingresos_qs = (reparaciones
                  .annotate(m=TruncMonth('fecha_ingreso'))
                  .values('m')
                  .annotate(
                      total=Sum('servicio__costo'),
                      cantidad=Count('id')
                  )
                  .order_by('m'))
    
    # Verificar si hay datos para exportar
    if not ingresos_qs:
        return JsonResponse({'error': 'No hay datos para exportar'}, status=400)
    
    # Procesar datos
    meses = [item['m'].strftime('%b %Y') if item['m'] else '' for item in ingresos_qs]
    ingresos = [float(item['total']) if item['total'] is not None else 0.0 for item in ingresos_qs]
    cantidades = [item['cantidad'] for item in ingresos_qs]

    # Intentar con xlsxwriter (preferido para mejor rendimiento y formato)
    try:
        import xlsxwriter
        output = BytesIO()
        workbook = xlsxwriter.Workbook(output, {'in_memory': True})
        worksheet = workbook.add_worksheet('Ingresos')

        # Estilos
        header_format = workbook.add_format({
            'bold': True,
            'bg_color': '#4F81BD',  # Azul corporativo
            'font_color': 'white',
            'align': 'center',
            'valign': 'vcenter',
            'border': 1,
            'font_size': 12
        })

        # Formato para filas pares (gris claro)
        even_row_format = workbook.add_format({
            'bg_color': '#F2F2F2',  # Gris muy claro
            'border': 1,
            'font_size': 11
        })

        # Formato para filas impares (blanco)
        odd_row_format = workbook.add_format({
            'bg_color': '#FFFFFF',  # Blanco
            'border': 1,
            'font_size': 11
        })

        # Formato para moneda
        money_format = workbook.add_format({
            'num_format': '$#,##0.00',
            'border': 1,
            'font_size': 11
        })

        # Formato para celdas de texto
        text_format = workbook.add_format({
            'border': 1,
            'font_size': 11
        })

        # Ancho de columnas
        worksheet.set_column('A:A', 20)  # Mes
        worksheet.set_column('B:B', 25)  # Cantidad
        worksheet.set_column('C:C', 25)  # Ingresos

        # Escribir encabezados
        headers = ['Mes', 'Cantidad de Reparaciones', 'Ingresos Totales']
        for col_num, header in enumerate(headers):
            worksheet.write(0, col_num, header, header_format)

        # Escribir datos con formato de filas alternadas
        for row_num, (mes, cantidad, ingreso) in enumerate(zip(meses, cantidades, ingresos), start=1):
            # Alternar entre formatos de fila
            row_format = even_row_format if row_num % 2 == 0 else odd_row_format
            
            # Escribir datos
            worksheet.write(row_num, 0, mes, text_format if row_num % 2 == 1 else text_format)
            worksheet.write_number(row_num, 1, cantidad, row_format)
            worksheet.write_number(row_num, 2, ingreso, money_format if ingreso == 0 else money_format)

        # Crear gráfico
        chart = workbook.add_chart({'type': 'column'})
        last_row = len(meses)
        chart.add_series({
            'name':       'Ingresos',
            'categories': ['Ingresos', 1, 0, last_row, 0],
            'values':     ['Ingresos', 1, 2, last_row, 2],
            'fill':       {'color': '#4F81BD'},  # Mismo azul que el encabezado
            'border':     {'color': '#4F81BD'}
        })
        
        chart.set_title({'name': 'Ingresos por Mes'})
        chart.set_x_axis({'name': 'Mes'})
        chart.set_y_axis({
            'name': 'Ingresos ($)',
            'num_format': '$#,##0'
        })
        chart.set_legend({'position': 'none'})  # Ocultar leyenda ya que solo hay una serie
        
        # Insertar gráfico en la hoja
        worksheet.insert_chart('E2', chart, {
            'x_offset': 25, 
            'y_offset': 10,
            'x_scale': 1.5,
            'y_scale': 1.5
        })

        # Añadir filtros si se aplicaron fechas
        if fecha_desde_str or fecha_hasta_str:
            filtro_text = "Filtro aplicado: "
            if fecha_desde_str:
                filtro_text += f"Desde {fecha_desde_str}"
            if fecha_hasta_str:
                if fecha_desde_str:
                    filtro_text += " - "
                filtro_text += f"Hasta {fecha_hasta_str}"
            
            # Escribir texto del filtro
            worksheet.merge_range(f'E{last_row + 3}:H{last_row + 3}', filtro_text, workbook.add_format({
                'italic': True,
                'font_color': '#666666',
                'font_size': 10
            }))

        # Cerrar el libro
        workbook.close()
        output.seek(0)

        # Crear la respuesta
        filename = f"reporte_ingresos_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        response = HttpResponse(
            output,
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename={filename}'
        return response

    except ImportError:
        # Si xlsxwriter no está disponible, intentar con openpyxl
        try:
            from openpyxl import Workbook
            from openpyxl.chart import BarChart, Reference
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
            from openpyxl.utils import get_column_letter

            output = BytesIO()
            wb = Workbook()
            ws = wb.active
            ws.title = "Ingresos"

            # Estilos
            header_fill = PatternFill(start_color='4F81BD', end_color='4F81BD', fill_type='solid')
            header_font = Font(color='FFFFFF', bold=True, size=12)
            border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )

            # Encabezados
            headers = ['Mes', 'Cantidad de Reparaciones', 'Ingresos Totales']
            for col_num, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col_num, value=header)
                cell.fill = header_fill
                cell.font = header_font
                cell.border = border
                cell.alignment = Alignment(horizontal='center', vertical='center')

            # Datos
            for row_num, (mes, cantidad, ingreso) in enumerate(zip(meses, cantidades, ingresos), start=2):
                # Alternar colores de fila
                if row_num % 2 == 0:
                    fill = PatternFill(start_color='F2F2F2', end_color='F2F2F2', fill_type='solid')
                else:
                    fill = PatternFill(start_color='FFFFFF', end_color='FFFFFF', fill_type='solid')
                
                # Escribir celdas
                ws.cell(row=row_num, column=1, value=mes).fill = fill
                ws.cell(row=row_num, column=2, value=cantidad).fill = fill
                cell_ingreso = ws.cell(row=row_num, column=3, value=ingreso)
                cell_ingreso.number_format = '$#,##0.00'
                cell_ingreso.fill = fill
                
                # Aplicar bordes
                for col in range(1, 4):
                    ws.cell(row=row_num, column=col).border = border

            # Ajustar ancho de columnas
            for col in ws.columns:
                max_length = 0
                column = col[0].column_letter
                for cell in col:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = (max_length + 2)
                ws.column_dimensions[column].width = adjusted_width

            # Crear gráfico
            chart = BarChart()
            chart.title = 'Ingresos por Mes'
            chart.x_axis.title = 'Mes'
            chart.y_axis.title = 'Ingresos ($)'

            data = Reference(ws, min_col=3, min_row=1, max_row=len(meses)+1, max_col=3)
            cats = Reference(ws, min_col=1, min_row=2, max_row=len(meses)+1)
            chart.add_data(data, titles_from_data=True)
            chart.set_categories(cats)
            
            # Personalizar el gráfico
            chart.series[0].graphicalProperties.solidFill = '4F81BD'  # Mismo azul que el encabezado
            chart.series[0].graphicalProperties.line.solidFill = '4F81BD'
            
            # Añadir el gráfico a la hoja
            ws.add_chart(chart, 'E2')

            # Añadir filtros si se aplicaron fechas
            if fecha_desde_str or fecha_hasta_str:
                filtro_text = "Filtro aplicado: "
                if fecha_desde_str:
                    filtro_text += f"Desde {fecha_desde_str}"
                if fecha_hasta_str:
                    if fecha_desde_str:
                        filtro_text += " - "
                    filtro_text += f"Hasta {fecha_hasta_str}"
                
                # Escribir texto del filtro
                ws.merge_cells(f'E{len(meses)+3}:H{len(meses)+3}')
                cell_filtro = ws.cell(row=len(meses)+3, column=5, value=filtro_text)
                cell_filtro.font = Font(italic=True, color='666666', size=10)

            # Guardar el libro
            wb.save(output)
            output.seek(0)

            # Crear la respuesta
            filename = f"reporte_ingresos_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
            response = HttpResponse(
                output,
                content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            )
            response['Content-Disposition'] = f'attachment; filename={filename}'
            return response

        except ImportError:
            # Si no hay bibliotecas de Excel, usar CSV como último recurso
            response = HttpResponse(content_type='text/csv')
            response['Content-Disposition'] = f'attachment; filename=reporte_ingresos_{datetime.now().strftime("%Y%m%d_%H%M%S")}.csv'

            writer = csv.writer(response)
            writer.writerow(['Mes', 'Cantidad Reparaciones', 'Ingresos'])
            for mes, cantidad, ingreso in zip(meses, cantidades, ingresos):
                writer.writerow([mes, cantidad, f'${ingreso:,.2f}'])

            return response

# ========== CLIENTES (CRUD con templates) ==========

@login_required
def clientes_lista(request):
    clientes = Cliente.objects.prefetch_related('vehiculos').order_by('nombre', 'apellido')
    return render(request, 'clientes_lista.html', {'clientes': clientes})


@login_required
@transaction.atomic
def clientes_crear(request):
    VehiculoFormSet = inlineformset_factory(Cliente, Vehiculo, fields=['marca', 'modelo', 'año', 'placa'], extra=1, can_delete=True)
    if request.method == 'POST':
        form = ClienteForm(request.POST)
        formset = VehiculoFormSet(request.POST, prefix='vehiculos')
        if form.is_valid() and formset.is_valid():
            cliente = form.save()
            formset.instance = cliente
            formset.save()
            messages.success(request, 'Cliente creado correctamente.')
            return redirect('clientes-lista')
    else:
        form = ClienteForm()
        formset = VehiculoFormSet(prefix='vehiculos')
    return render(request, 'clientes_form.html', {'form': form, 'formset': formset, 'accion': 'Crear'})


@login_required
@transaction.atomic
def clientes_editar(request, pk):
    cliente = get_object_or_404(Cliente, pk=pk)
    VehiculoFormSet = inlineformset_factory(Cliente, Vehiculo, fields=['marca', 'modelo', 'año', 'placa'], extra=0, can_delete=True)
    if request.method == 'POST':
        form = ClienteForm(request.POST, instance=cliente)
        formset = VehiculoFormSet(request.POST, instance=cliente, prefix='vehiculos')
        if form.is_valid() and formset.is_valid():
            form.save()
            formset.save()
            messages.success(request, 'Cliente actualizado correctamente.')
            return redirect('clientes-lista')
    else:
        form = ClienteForm(instance=cliente)
        formset = VehiculoFormSet(instance=cliente, prefix='vehiculos')
    return render(request, 'clientes_form.html', {'form': form, 'formset': formset, 'accion': 'Editar'})


@login_required
def clientes_eliminar(request, pk):
    cliente = get_object_or_404(Cliente, pk=pk)
    if request.method == 'POST':
        cliente.delete()
        messages.success(request, 'Cliente eliminado correctamente.')
        return redirect('clientes-lista')
    return render(request, 'clientes_confirm_delete.html', {'cliente': cliente})


# ========== EMPLEADOS (CRUD con templates) ==========

@login_required
def empleados_lista(request):
    empleados = Empleado.objects.all().order_by('nombre')
    return render(request, 'empleados_lista.html', {'empleados': empleados})


@login_required
def empleados_crear(request):
    if request.method == 'POST':
        form = EmpleadoForm(request.POST)
        if form.is_valid():
            form.save()
            messages.success(request, 'Empleado creado correctamente.')
            return redirect('empleados-lista')
    else:
        form = EmpleadoForm()
    return render(request, 'empleados_form.html', {'form': form, 'accion': 'Crear'})


@login_required
def empleados_editar(request, pk):
    empleado = get_object_or_404(Empleado, pk=pk)
    if request.method == 'POST':
        form = EmpleadoForm(request.POST, instance=empleado)
        if form.is_valid():
            form.save()
            messages.success(request, 'Empleado actualizado correctamente.')
            return redirect('empleados-lista')
    else:
        form = EmpleadoForm(instance=empleado)
    return render(request, 'empleados_form.html', {'form': form, 'accion': 'Editar'})


@login_required
def empleados_eliminar(request, pk):
    empleado = get_object_or_404(Empleado, pk=pk)
    if request.method == 'POST':
        empleado.delete()
        messages.success(request, 'Empleado eliminado correctamente.')
        return redirect('empleados-lista')
    return render(request, 'empleados_confirm_delete.html', {'empleado': empleado})


# ========== SERVICIOS (CRUD con templates) ==========

@login_required
def servicios_lista(request):
    servicios = Servicio.objects.all().order_by('nombre_servicio')
    return render(request, 'servicios_lista.html', {'servicios': servicios})


@login_required
def servicios_crear(request):
    if request.method == 'POST':
        form = ServicioForm(request.POST)
        if form.is_valid():
            form.save()
            messages.success(request, 'Servicio creado correctamente.')
            return redirect('servicios-lista')
    else:
        form = ServicioForm()
    return render(request, 'servicios_form.html', {'form': form, 'accion': 'Crear'})


@login_required
def servicios_editar(request, pk):
    servicio = get_object_or_404(Servicio, pk=pk)
    if request.method == 'POST':
        form = ServicioForm(request.POST, instance=servicio)
        if form.is_valid():
            form.save()
            messages.success(request, 'Servicio actualizado correctamente.')
            return redirect('servicios-lista')
    else:
        form = ServicioForm(instance=servicio)
    return render(request, 'servicios_form.html', {'form': form, 'accion': 'Editar'})


@login_required
def servicios_eliminar(request, pk):
    servicio = get_object_or_404(Servicio, pk=pk)
    if request.method == 'POST':
        servicio.delete()
        messages.success(request, 'Servicio eliminado correctamente.')
        return redirect('servicios-lista')
    return render(request, 'servicios_confirm_delete.html', {'servicio': servicio})


# ========== VEHÍCULOS (crear/editar/eliminar con template) ==========

@login_required
def vehiculo_agregar(request, cliente_id=None):
    titulo = 'Agregar Vehículo'
    cliente = None
    if cliente_id:
        cliente = get_object_or_404(Cliente, pk=cliente_id)
    if request.method == 'POST':
        form = VehiculoForm(request.POST)
        if form.is_valid():
            form.save()
            messages.success(request, 'Vehículo agregado correctamente.')
            return redirect('vehiculos-lista')
    else:
        form = VehiculoForm(initial={'cliente': cliente.id} if cliente else None)
    return render(request, 'gestion/vehiculo_form.html', {'form': form, 'titulo': titulo, 'cliente': cliente})


@login_required
def vehiculo_editar(request, pk):
    vehiculo = get_object_or_404(Vehiculo, pk=pk)
    titulo = 'Editar Vehículo'
    if request.method == 'POST':
        form = VehiculoForm(request.POST, instance=vehiculo)
        if form.is_valid():
            form.save()
            messages.success(request, 'Vehículo actualizado correctamente.')
            return redirect('vehiculos-lista')
    else:
        form = VehiculoForm(instance=vehiculo)
    return render(request, 'gestion/vehiculo_form.html', {'form': form, 'titulo': titulo, 'cliente': vehiculo.cliente})


@login_required
def vehiculo_eliminar(request, pk):
    vehiculo = get_object_or_404(Vehiculo, pk=pk)
    if request.method == 'POST':
        vehiculo.delete()
        messages.success(request, 'Vehículo eliminado correctamente.')
        return redirect('vehiculos-lista')
    return render(request, 'vehiculos_confirm_delete.html', {'vehiculo': vehiculo})

# ========== API VIEWS Y VIEWSETS ==========

# Cliente
class ClienteListCreate(generics.ListCreateAPIView):
    queryset = Cliente.objects.all()
    serializer_class = ClienteSerializer


class ClienteRetrieveUpdateDestroy(generics.RetrieveUpdateDestroyAPIView):
    queryset = Cliente.objects.all()
    serializer_class = ClienteSerializer


# Empleado
class EmpleadoListCreate(generics.ListCreateAPIView):
    queryset = Empleado.objects.all()
    serializer_class = EmpleadoSerializer


class EmpleadoRetrieveUpdateDestroy(generics.RetrieveUpdateDestroyAPIView):
    queryset = Empleado.objects.all()
    serializer_class = EmpleadoSerializer


# Servicio
class ServicioListCreate(generics.ListCreateAPIView):
    queryset = Servicio.objects.all()
    serializer_class = ServicioSerializer


class ServicioRetrieveUpdateDestroy(generics.RetrieveUpdateDestroyAPIView):
    queryset = Servicio.objects.all()
    serializer_class = ServicioSerializer


# Vehiculo
class VehiculoListCreate(generics.ListCreateAPIView):
    queryset = Vehiculo.objects.all()
    serializer_class = VehiculoSerializer


class VehiculoRetrieveUpdateDestroy(generics.RetrieveUpdateDestroyAPIView):
    queryset = Vehiculo.objects.all()
    serializer_class = VehiculoSerializer


# Reparacion
class ReparacionListCreate(generics.ListCreateAPIView):
    queryset = Reparacion.objects.all()
    serializer_class = ReparacionSerializer


class ReparacionRetrieveUpdateDestroy(generics.RetrieveUpdateDestroyAPIView):
    queryset = Reparacion.objects.all()
    serializer_class = ReparacionSerializer


# Agenda (citas)
class AgendaViewSet(viewsets.ModelViewSet):
    queryset = Agenda.objects.all()
    serializer_class = AgendaSerializer

    def create(self, request, *args, **kwargs):
        serializer = self.get_serializer(data=request.data)
        serializer.is_valid(raise_exception=True)
        try:
            cita = Agenda().programarCita(
                cliente=serializer.validated_data['cliente'],
                servicio=serializer.validated_data['servicio'],
                fecha=serializer.validated_data['fecha'],
                hora=serializer.validated_data['hora']
            )
        except ValidationError as e:
            return Response({'error': e.message}, status=status.HTTP_400_BAD_REQUEST)
        output_serializer = self.get_serializer(cita)
        return Response(output_serializer.data, status=status.HTTP_201_CREATED)


# Registro (historial)
class RegistroViewSet(viewsets.ModelViewSet):
    queryset = Registro.objects.all()
    serializer_class = RegistroSerializer

    def create(self, request, *args, **kwargs):
        serializer = self.get_serializer(data=request.data)
        serializer.is_valid(raise_exception=True)
        try:
            registro = Registro().crearRegistro(
                cliente=serializer.validated_data['cliente'],
                empleado=serializer.validated_data['empleado'],
                servicio=serializer.validated_data['servicio'],
                fecha=serializer.validated_data.get('fecha', None)
            )
        except ValidationError as e:
            return Response({'error': e.message}, status=status.HTTP_400_BAD_REQUEST)
        output_serializer = self.get_serializer(registro)
        return Response(output_serializer.data, status=status.HTTP_201_CREATED)


# ========== VISTAS DE CITAS ==========

@login_required
def lista_citas(request):
    """Vista para listar todas las citas"""
    citas = Agenda.objects.select_related('cliente', 'servicio').order_by('-fecha', '-hora')
    return render(request, 'gestion/citas/lista_citas.html', {
        'titulo': 'Lista de Citas',
        'citas': citas
    })

@login_required
def crear_cita(request):
    """Vista para crear una nueva cita"""
    if request.method == 'POST':
        form = CitaForm(request.POST)
        if form.is_valid():
            try:
                cita = form.save(commit=False)
                # Validar que la fecha no sea pasada
                if cita.fecha < timezone.now().date():
                    messages.error(request, 'No se pueden agendar citas en fechas pasadas.')
                # Validar que no haya otra cita a la misma hora
                elif Agenda.objects.filter(fecha=cita.fecha, hora=cita.hora).exists():
                    messages.error(request, 'Ya existe una cita programada para esa fecha y hora.')
                else:
                    cita.save()
                    messages.success(request, 'Cita creada exitosamente.')
                    return redirect('lista_citas')
            except Exception as e:
                messages.error(request, f'Error al crear la cita: {str(e)}')
    else:
        form = CitaForm()
    
    return render(request, 'gestion/citas/crear_cita.html', {
        'titulo': 'Nueva Cita',
        'form': form
    })

@login_required
def editar_cita(request, pk):
    """Vista para editar una cita existente"""
    cita = get_object_or_404(Agenda, pk=pk)
    
    if request.method == 'POST':
        form = CitaForm(request.POST, instance=cita)
        if form.is_valid():
            try:
                cita_editada = form.save(commit=False)
                # Validar que la fecha no sea pasada
                if cita_editada.fecha < timezone.now().date():
                    messages.error(request, 'No se pueden agendar citas en fechas pasadas.')
                # Validar que no haya otra cita a la misma hora (excluyendo la actual)
                elif Agenda.objects.filter(fecha=cita_editada.fecha, hora=cita_editada.hora).exclude(pk=pk).exists():
                    messages.error(request, 'Ya existe otra cita programada para esa fecha y hora.')
                else:
                    cita_editada.save()
                    messages.success(request, 'Cita actualizada exitosamente.')
                    return redirect('lista_citas')
            except Exception as e:
                messages.error(request, f'Error al actualizar la cita: {str(e)}')
    else:
        form = CitaForm(instance=cita)
    
    return render(request, 'gestion/citas/editar_cita.html', {
        'titulo': 'Editar Cita',
        'form': form,
        'cita': cita
    })

@login_required
def eliminar_cita(request, pk):
    """Vista para eliminar una cita"""
    cita = get_object_or_404(Agenda, pk=pk)
    
    if request.method == 'POST':
        try:
            cita.delete()
            messages.success(request, 'Cita eliminada exitosamente.')
        except Exception as e:
            messages.error(request, f'Error al eliminar la cita: {str(e)}')
        return redirect('lista_citas')
    
    return render(request, 'gestion/citas/eliminar_cita.html', {
        'titulo': 'Eliminar Cita',
        'cita': cita
    })

@login_required
@api_view(['GET'])
def obtener_horas_disponibles(request, fecha):
    """API para obtener las horas disponibles en una fecha específica"""
    try:
        # Convertir la fecha de string a objeto date
        fecha_obj = datetime.strptime(fecha, '%Y-%m-%d').date()
        
        # Obtener las horas ocupadas para la fecha dada
        horas_ocupadas = list(Agenda.objects.filter(
            fecha=fecha_obj
        ).values_list('hora', flat=True))
        
        # Generar lista de horas disponibles (de 8:00 a 17:00, cada 30 minutos)
        horas_disponibles = []
        hora_actual = dt_time(8, 0)  # Empieza a las 8:00 AM
        hora_fin = dt_time(17, 0)    # Termina a las 5:00 PM
        
        while hora_actual <= hora_fin:
            # Verificar si la hora actual no está ocupada
            if hora_actual not in horas_ocupadas:
                horas_disponibles.append(hora_actual.strftime('%H:%M'))
            
            # Añadir 30 minutos
            hora_actual = (datetime.combine(datetime.today(), hora_actual) + 
                          timedelta(minutes=30)).time()
        
        return JsonResponse({'horas_disponibles': horas_disponibles})
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=400)

# ========== VISTAS BASADAS EN CLASES ==========

class VehiculoListView(ListView):
    """
    Vista basada en clase para mostrar la lista de vehículos con búsqueda y paginación.
    """
    model = Vehiculo
    template_name = 'gestion/vehiculo_list.html'
    context_object_name = 'vehiculos'
    paginate_by = 10
    
    def get_queryset(self):
        queryset = super().get_queryset().select_related('cliente')
        busqueda = self.request.GET.get('q', '').strip()
        
        if busqueda:
            # Búsqueda por placa, marca, modelo o nombre del cliente
            queryset = queryset.filter(
                Q(placa__icontains=busqueda) |
                Q(marca__icontains=busqueda) |
                Q(modelo__icontains=busqueda) |
                Q(cliente__nombre__icontains=busqueda) |
                Q(cliente__apellido__icontains=busqueda)
            )
        
        return queryset.order_by('-id')
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['titulo'] = 'Lista de Vehículos'
        context['busqueda'] = self.request.GET.get('q', '')
        return context


class APIAuthenticationMixin:
    """
    Mixin para agregar autenticación a las vistas de API.
    Asegura que solo los usuarios autenticados puedan acceder a las vistas.
    """
    authentication_classes = [SessionAuthentication]
    permission_classes = [IsAuthenticated]

# Aplicar autenticación a todas las vistas de API
ClienteListCreate.authentication_classes = [SessionAuthentication]
ClienteListCreate.permission_classes = [IsAuthenticated]
ClienteRetrieveUpdateDestroy.authentication_classes = [SessionAuthentication]
ClienteRetrieveUpdateDestroy.permission_classes = [IsAuthenticated]

EmpleadoListCreate.authentication_classes = [SessionAuthentication]
EmpleadoListCreate.permission_classes = [IsAuthenticated]
EmpleadoRetrieveUpdateDestroy.authentication_classes = [SessionAuthentication]
EmpleadoRetrieveUpdateDestroy.permission_classes = [IsAuthenticated]

ServicioListCreate.authentication_classes = [SessionAuthentication]
ServicioListCreate.permission_classes = [IsAuthenticated]
ServicioRetrieveUpdateDestroy.authentication_classes = [SessionAuthentication]
ServicioRetrieveUpdateDestroy.permission_classes = [IsAuthenticated]

VehiculoListCreate.authentication_classes = [SessionAuthentication]
VehiculoListCreate.permission_classes = [IsAuthenticated]
VehiculoRetrieveUpdateDestroy.authentication_classes = [SessionAuthentication]
VehiculoRetrieveUpdateDestroy.permission_classes = [IsAuthenticated]

ReparacionListCreate.authentication_classes = [SessionAuthentication]
ReparacionListCreate.permission_classes = [IsAuthenticated]
ReparacionRetrieveUpdateDestroy.authentication_classes = [SessionAuthentication]
ReparacionRetrieveUpdateDestroy.permission_classes = [IsAuthenticated]

AgendaViewSet.authentication_classes = [SessionAuthentication]
AgendaViewSet.permission_classes = [IsAuthenticated]
RegistroViewSet.authentication_classes = [SessionAuthentication]
RegistroViewSet.permission_classes = [IsAuthenticated]
